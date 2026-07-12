#!/usr/bin/env python3
"""
VOX Extraction -> mockVistaData.js converter.
Reads the showtimes extraction workbook and emits src/mockVistaData.js in the
exact Vista-shaped structures the widget consumes. Re-run any time you refresh
the extraction.

Usage: python3 convert_extraction.py <extraction.xlsx> <output mockVistaData.js>
"""
import sys, json, hashlib
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/VOX_Cinemas_Showtimes_08-15Jul_1.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "src/mockVistaData.js"

# Demo scope: flagship cinemas (edit this list to change scope)
FLAGSHIPS = ["Mall of the Emirates", "The Galleria Al Maryah Island", "City Centre Mirdif"]

# Deterministic brand-adjacent tint per movie (fallback if poster fails to load)
PALETTE = [("#63418D","#B6186C"),("#7A5A2E","#C79A4B"),("#2E6A5A","#57C79A"),
           ("#8D2E3A","#D9556B"),("#3A3A8D","#6B6BD9"),("#2E5A7A","#4BA3C7"),
           ("#6A2E7A","#B44BC7"),("#7A2E2E","#C74B4B")]
def tint(code):
    return PALETTE[int(hashlib.md5(code.encode()).hexdigest(), 16) % len(PALETTE)]

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

# --- Movie Catalog -> film metadata ---------------------------------------
cat = {}
for r in wb["Movie Catalog"].iter_rows(min_row=2, values_only=True):
    if not r or not r[0]: continue
    title, code, rating, lang, total, genres, subs, released, synopsis = (list(r)+[None]*9)[:9]
    cat[title] = dict(code=str(code or ""), rating=str(rating or ""), lang=str(lang or ""),
                      genres=str(genres or ""), synopsis=str(synopsis or "")[:220])

# --- Showtimes -> cinemas + sessions (flagships only) ----------------------
cinemas, sessions, films_at = {}, [], {}
sid = 10000
for r in wb["Showtimes"].iter_rows(min_row=2, values_only=True):
    if not r or not r[0]: continue
    date, day, movie, fcode, rating, lang, ccode, cinema, exp, time, slot, status = (list(r)+[None]*12)[:12]
    if cinema not in FLAGSHIPS: continue
    ccode = str(ccode)
    cinemas.setdefault(ccode, dict(ID=ccode, Name=f"VOX — {cinema}", City="UAE", CurrencyCode="AED"))
    meta = cat.get(movie, {})
    code = str(fcode or meta.get("code") or "")
    films_at.setdefault((ccode, code), dict(
        ScheduledFilmId=code, CinemaId=ccode, Title=str(movie),
        Rating=str(rating or meta.get("rating","")), Language=str(lang or meta.get("lang","")),
        genre=(meta.get("genres","").split(",")[0].strip() or "Film"),
        Synopsis=meta.get("synopsis",""),
        posterUrl=f"https://assets.voxcinemas.com/posters/P_{code}.jpg",
        tint=list(tint(code)),
    ))
    sid += 1
    # deterministic seats-available per session (looks live, stable per run)
    seats = 40 + int(hashlib.md5(f"{ccode}{code}{date}{time}".encode()).hexdigest(), 16) % 140
    sessions.append(dict(
        CinemaId=ccode, ScheduledFilmId=code, SessionId=str(sid),
        Showtime=f"{str(date)[:10]}T{str(time)}:00", ScreenName=str(exp or "Screen"),
        SeatsAvailable=seats, SessionAttributesNames=[str(exp or "2D")],
    ))

films = list(films_at.values())
cin_list = list(cinemas.values())
dates = sorted({s["Showtime"][:10] for s in sessions})

js = f"""// ============================================================================
//  REAL VOX DATA — generated from the showtimes extraction (uae.voxcinemas.com)
//  Coverage: {dates[0]} to {dates[-1]} | Cinemas: {len(cin_list)} flagships | Films: {len(films)} | Sessions: {len(sessions)}
//  Regenerate with convert_extraction.py after a fresh extraction.
// ============================================================================

export const DATA_DATES = {json.dumps(dates)};

export const CINEMAS = {json.dumps(cin_list, ensure_ascii=False, indent=2)};

export const FILMS = {json.dumps(films, ensure_ascii=False, indent=2)};

export const SESSIONS = {json.dumps(sessions, ensure_ascii=False)};

// Deterministic seat plan in the real Vista seat-plan shape.
export function seatPlan(seed = 7) {{
  let n = seed || 1;
  const rnd = () => {{ n = (n * 9301 + 49297) % 233280; return n / 233280; }};
  const rows = "ABCDEFGH".split("").map((name, ri) => ({{
    RowIndexZeroBased: ri,
    PhysicalName: name,
    Seats: Array.from({{ length: 12 }}, (_, ci) => ({{
      Position: {{ AreaNumber: 1, RowIndex: ri, ColumnIndex: ci }},
      Id: String(ci + 1),
      Status: rnd() < 0.22 ? 1 : 0,
      SeatStyle: 0,
      areaCategoryCode: ri >= 5 ? "0000000001" : "0000000002",
    }})),
  }}));
  return {{
    SeatLayoutData: {{
      Areas: [{{ AreaCategoryCode: "0000000002", Description: "REGULAR", Rows: rows, RowCount: 8, ColumnCount: 12 }}],
      AreaCategories: [
        {{ AreaCategoryCode: "0000000002", Name: "REGULAR" }},
        {{ AreaCategoryCode: "0000000001", Name: "PREMIUM" }},
      ],
    }},
    ResponseCode: 0,
    ErrorDescription: null,
  }};
}}

export const BOOKING = {{
  BookingId: "WL59LFJ", BookingNumber: 8608,
  FilmTitle: {json.dumps(films[0]["Title"] if films else "Movie")}, Showtime: "{dates[0]}T18:40:00",
  Seats: ["C5", "C6"], TotalValueCents: 12600, ScreenName: "MAX",
}};
"""
with open(OUT, "w", encoding="utf-8") as f:
    f.write(js)
print(f"Wrote {OUT}: {len(cin_list)} cinemas, {len(films)} films, {len(sessions)} sessions, dates {dates[0]}..{dates[-1]}")
